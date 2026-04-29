import pandas as pd
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes
from datetime import datetime
import nest_asyncio

nest_asyncio.apply()

TOKEN = "7119344534:AAFJP-0BM9OvIoo_bw2RDB5nfm0HVGBKKxQ"
EXCEL = "RECORDATORIOS.xlsm"


# Leer Excel
def leer_excel():
    df = pd.read_excel(EXCEL, header=None)

    # Buscar fila donde esté "Alertas"
    for i, row in df.iterrows():
        if row.astype(str).str.contains("Alertas", case=False).any():
            df.columns = df.iloc[i]   # usar esa fila como encabezado
            df = df[i+1:]             # datos después del encabezado
            df = df.reset_index(drop=True)
            break

    print("COLUMNAS REALES:", df.columns)
    return df

#############stratr ver funciones 
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mensaje = (
        "🤖 *BOT DE RECORDATORIOS*\n\n"

        "📌 *COMANDOS DISPONIBLES:*\n\n"

        "📄 /todo\n"
        "➡️ Muestra todos los registros del Excel\n\n"

        "👀 /ver\n"
        "➡️ Ver facturas activas\n\n"

        "🚨 /vencidas\n"
        "➡️ Muestra facturas vencidas\n\n"

        "🔎 /buscar nombre\n"
        "➡️ Buscar por cliente\n"
        "Ejemplo: /buscar juan\n\n"

        "🎵 /tipo servicio\n"
        "➡️ Buscar por tipo de cuenta\n"
        "Ejemplo: /tipo spotify\n\n"

        "🔑 /cuentas\n"
        "➡️ Ver todas las cuentas\n\n"

        "📭 /vacios\n"
        "➡️ Ver registros vacíos\n\n"

        "➕ /agregar nombre | fecha | dias | servicio | cuenta\n"
        "➡️ Agregar nuevo cliente\n"
        "Ejemplo:\n"
        "/agregar Juan Perez | 2025-05-01 | 30 | Spotify | correo@gmail.com:123\n\n"

        "⚡ Usa los comandos correctamente para evitar errores"
    )

    await update.message.reply_text(mensaje, parse_mode="Markdown")
# /todo
async def todo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = leer_excel()

    mensajes = []
    actual = ""

    for i, row in df.iterrows():
        if row.notna().any():
            fila_texto = " | ".join([str(x) for x in row if pd.notna(x)]) + "\n"

            if len(actual) + len(fila_texto) > 4000:
                mensajes.append(actual)
                actual = ""

            actual += fila_texto

    if actual:
        mensajes.append(actual)

    if not mensajes:
        await update.message.reply_text("No hay datos")
        return

    for parte in mensajes:
        await update.message.reply_text(parte)


# /ver
async def ver(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = leer_excel()

    mensaje = ""
    for i, row in df.iterrows():
        if str(row.iloc[6]).upper() != "VACIO":
            mensaje += f"{row.iloc[0]} | {row.iloc[1]} | {row.iloc[6]}\n"

    await update.message.reply_text(mensaje or "No hay datos activos")


# /vencidas (VERSIÓN SEGURA)
async def vencidas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = leer_excel()

    vencidas_lista = []

    # Filtrar vencidas
    for i, row in df.iterrows():
        alerta = str(row["Alertas"]).upper()

        if "VENCIDO" in alerta:
            vencidas_lista.append(row)

    total = len(vencidas_lista)

    # 📊 RESUMEN
    if total == 0:
        await update.message.reply_text("✅ No hay facturas vencidas")
        return
    else:
        await update.message.reply_text(
            f"🚨 *FACTURAS VENCIDAS*\n\n📊 Total: {total}",
            parse_mode="Markdown"
        )

    # 📩 DETALLE UNA POR UNA
    for row in vencidas_lista:

        nombre = row["Cliente"]
        fecha = str(row["Fecha Emisión"]).split(" ")[0]
        dias = row["Días a vencer"]
        vencimiento = str(row["Fecha Vencimiento"]).split(" ")[0]
        estado = row["Alertas"]
        servicio = str(row["TIPO DE CUNTA"])
        cuenta = str(row["HGFHF"])

        # 🎯 Emoji según servicio
        if "spotify" in servicio.lower():
            emoji_servicio = "🎵"
        elif "hbo" in servicio.lower():
            emoji_servicio = "🎬"
        elif "netflix" in servicio.lower():
            emoji_servicio = "🍿"
        else:
            emoji_servicio = "📦"

        # 🚨 Emoji estado
        if "VENCIDO" in estado.upper():
            emoji_estado = "❌"
        else:
            emoji_estado = "⚠️"

        mensaje = (
            f"👤 *NOMBRE:* {nombre}\n"
            f"📅 *FECHA:* {fecha}\n"
            f"⏳ *DIAS:* {dias}\n"
            f"📆 *VENCIMIENTO:* {vencimiento}\n"
            f"{emoji_estado} *ESTADO:* {estado}\n"
            f"{emoji_servicio} *SERVICIO:* {servicio}\n"
            f"🔑 *CUENTA:* `{cuenta}`"
        )

        await update.message.reply_text(mensaje, parse_mode="Markdown")

########################## /buscar cliente 

async def buscar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = leer_excel()

    # Obtener texto que escribe el usuario
    if not context.args:
        await update.message.reply_text("❗ Usa: /buscar nombre")
        return

    texto = " ".join(context.args).lower()

    resultados = []

    for i, row in df.iterrows():
        nombre = str(row["Cliente"]).lower()

        if texto in nombre:
            resultados.append(row)

    total = len(resultados)

    # 📊 RESUMEN
    if total == 0:
        await update.message.reply_text("❌ No se encontraron coincidencias")
        return
    else:
        await update.message.reply_text(
            f"🔎 *RESULTADOS DE BÚSQUEDA*\n\n👤 Nombre: {texto}\n📊 Coincidencias: {total}",
            parse_mode="Markdown"
        )

    # 📩 RESULTADOS
    for row in resultados:

        nombre = row["Cliente"]
        fecha = str(row["Fecha Emisión"]).split(" ")[0]
        dias = row["Días a vencer"]
        vencimiento = str(row["Fecha Vencimiento"]).split(" ")[0]
        estado = row["Alertas"]
        servicio = str(row["TIPO DE CUNTA"])
        cuenta = str(row["HGFHF"])

        # Emojis servicio
        if "spotify" in servicio.lower():
            emoji_servicio = "🎵"
        elif "hbo" in servicio.lower():
            emoji_servicio = "🎬"
        elif "netflix" in servicio.lower():
            emoji_servicio = "🍿"
        else:
            emoji_servicio = "📦"

        # Emoji estado
        if "VENCIDO" in str(estado).upper():
            emoji_estado = "❌"
        elif "FALTAN" in str(estado).upper():
            emoji_estado = "⏳"
        else:
            emoji_estado = "⚪"

        mensaje = (
            f"👤 *NOMBRE:* {nombre}\n"
            f"📅 *FECHA:* {fecha}\n"
            f"⏳ *DIAS:* {dias}\n"
            f"📆 *VENCIMIENTO:* {vencimiento}\n"
            f"{emoji_estado} *ESTADO:* {estado}\n"
            f"{emoji_servicio} *SERVICIO:* {servicio}\n"
            f"🔑 *CUENTA:* `{cuenta}`"
        )

        await update.message.reply_text(mensaje, parse_mode="Markdown")
        
# /cuentas
async def cuentas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = leer_excel()

    mensaje = ""
    for i, row in df.iterrows():
        if isinstance(row.iloc[7], str):
            mensaje += f"{row.iloc[0]} → {row.iloc[7]}\n"

    await update.message.reply_text(mensaje or "Sin cuentas")


# /vacios
async def vacios(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = leer_excel()

    mensaje = ""
    for i, row in df.iterrows():
        if str(row.iloc[6]).upper() == "VACIO":
            mensaje += f"{row.iloc[0]}\n"

    await update.message.reply_text(mensaje or "No hay vacíos")

##################### buscar cuentas 
async def tipo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = leer_excel()

    if not context.args:
        await update.message.reply_text("❗ Usa: /tipo spotify | hbo | netflix")
        return

    servicio_buscar = " ".join(context.args).lower()

    resultados = []

    for i, row in df.iterrows():
        servicio = str(row["TIPO DE CUNTA"]).lower()

        if servicio_buscar in servicio:
            resultados.append(row)

    total = len(resultados)

    if total == 0:
        await update.message.reply_text("❌ No se encontraron cuentas de ese tipo")
        return

    # 🎯 Emoji por servicio
    if "spotify" in servicio_buscar:
        emoji = "🎵"
    elif "hbo" in servicio_buscar:
        emoji = "🎬"
    elif "netflix" in servicio_buscar:
        emoji = "🍿"
    else:
        emoji = "📦"

    # 📊 RESUMEN
    await update.message.reply_text(
        f"{emoji} *TIPO DE CUENTA: {servicio_buscar.upper()}*\n\n📊 Total cuentas: {total}",
        parse_mode="Markdown"
    )

    # 📩 LISTA DE CUENTAS
    for row in resultados:
        cuenta = str(row["HGFHF"])
        nombre = row["Cliente"]

        mensaje = (
            f"{emoji} *SERVICIO:* {servicio_buscar.upper()}\n"
            f"👤 *CLIENTE:* {nombre}\n"
            f"🔑 *CUENTA:*\n`{cuenta}`"
        )

        await update.message.reply_text(mensaje, parse_mode="Markdown")


#################### agregar cliente 
from openpyxl import load_workbook

async def agregar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        texto = " ".join(context.args)

        if "|" not in texto:
            await update.message.reply_text(
                "❗ Formato:\n/agregar nombre | fecha | dias | servicio | cuenta"
            )
            return

        partes = [x.strip() for x in texto.split("|")]

        if len(partes) != 5:
            await update.message.reply_text("❗ Faltan datos")
            return

        nombre, fecha_str, dias, servicio, cuenta = partes

        fecha = pd.to_datetime(fecha_str)
        dias = int(dias)

        wb = load_workbook(EXCEL, keep_vba=True)
        ws = wb.active

        fila_vacia = None

        # 🔍 Buscar desde fila 4
        for row in range(4, ws.max_row + 1):
            if not ws.cell(row=row, column=3).value:  # Cliente vacío
                fila_vacia = row
                break

        if fila_vacia is None:
            fila_vacia = ws.max_row + 1

        # ✍️ Escribir datos (NO factura)
        ws.cell(row=fila_vacia, column=3, value=nombre)     # Cliente
        ws.cell(row=fila_vacia, column=4, value=fecha)      # Fecha
        ws.cell(row=fila_vacia, column=5, value=dias)       # Días
        ws.cell(row=fila_vacia, column=9, value=servicio)   # Tipo cuenta
        ws.cell(row=fila_vacia, column=10, value=cuenta)     # Cuenta

        wb.save(EXCEL)

        await update.message.reply_text(
            f"✅ Cliente agregado en fila {fila_vacia}\n👤 {nombre}"
        )

    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")
        
        
# Bot
app = ApplicationBuilder().token(TOKEN).build()

app.add_handler(CommandHandler("ver", ver))
app.add_handler(CommandHandler("cuentas", cuentas))
app.add_handler(CommandHandler("vacios", vacios))
app.add_handler(CommandHandler("todo", todo))
app.add_handler(CommandHandler("vencidas", vencidas))
app.add_handler(CommandHandler("buscar", buscar))
app.add_handler(CommandHandler("tipo", tipo))
app.add_handler(CommandHandler("agregar", agregar))
app.add_handler(CommandHandler("start", start))

print("Bot corriendo...")
app.run_polling()
